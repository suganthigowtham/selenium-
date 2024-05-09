
from openpyxl.reader.excel import load_workbook


class Webdata:

    def __init__(self):
        self.url = "https://opensource-demo.orangehrmlive.com/web/index.php/auth/login"
        self.dashboardURL = "https://opensource-demo.orangehrmlive.com/web/index.php/dashboard/index"
        self.fileName = "Data/DDT.xlsx"
        self.sheetName = "Sheet1"
        self.workbook = load_workbook(self.fileName)
        self.sheet = self.workbook[self.sheetName]

    def rowCount(self):
        """
        This method returns the maximum number of rows present in the sheet1
        :return: int
        """
        return self.sheet.max_row

    def readData(self, row, column):
        """
        This method will return the data present in that particular cell in sheet1


        :param row:
        :param column:
        :return:
        """
        return self.sheet.cell(row, column).value

    def writeData(self, row, column, data):
        """
        This method is used to write data in that particular cell in sheet1
        :param row:
        :param column:
        :param data:
        :return:
        """
        self.sheet.cell(row, column).value = data
        self.workbook.save(self.fileName)

